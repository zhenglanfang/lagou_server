#! /usr/bin/python
# coding:utf-8

from openpyxl.utils import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font,PatternFill,Color,Alignment


class WriteExcel(object):
    """
        openpyxl
        封装写excel的方法
    """
    font = Font(name='Consolas', size=15)
    title_font = Font(name='Consolas', size=15, bold=True)
    red_fill = PatternFill(fgColor='2981e2', fill_type='solid')

    @classmethod
    def write_title(cls, sheet, title, start_row=1, title_font=title_font):
        for index, value in enumerate(title):
            cell = sheet.cell(column=index + 1, row=start_row, value=value)
            cell.font = title_font
            # sheet.write(0, index, value, style)

    # 写入['V1','V2','V3']结构的数据
    @classmethod
    def write_excel_list(cls, sheet, title, data, wrap=False, font=font):
        cls.write_title(sheet, title)
        for i, item in enumerate(data):
            cell = sheet.cell(column=1, row=i + 2, value=item)
            cell.font = font
            cell.alignment = Alignment(wrap_text=wrap)
            # sheet.write(i+1, j, item[v], style)

    # 写入[{'k1':'v1'},...]结构的数据
    @classmethod
    def write_excel_dict(cls, sheet, title, data, wrap=False, font=font):
        cls.write_title(sheet, title)
        for i, item in enumerate(data):
            for j, v in enumerate(title):
                cell = sheet.cell(column=j + 1, row=i + 2, value=item[v])
                cell.font = font
                cell.alignment = Alignment(wrap_text=wrap)
                # sheet.write(i+1, j, item[v], style)

    # 设置excel自适应宽度
    def adgust_column_width(cls, data, worksheet):
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths):
            worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width